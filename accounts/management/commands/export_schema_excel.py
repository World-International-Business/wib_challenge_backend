import unicodedata
from django.apps import apps
from django.core.management.base import BaseCommand
from django.db import models
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def clean_text(value):
    """Nettoie les chaînes pour qu'elles soient compatibles Excel."""
    if not isinstance(value, str):
        return value
    # Normalise le texte (transforme les caractères typographiques)
    text = unicodedata.normalize("NFKC", value)
    # Remplace les guillemets/apostrophes typographiques par simples
    text = text.replace('’', "'").replace('“', '"').replace('”', '"')
    # Supprime les caractères non imprimables
    return ''.join(c for c in text if c.isprintable()).encode()


class Command(BaseCommand):
    help = "Exporte le schéma de la base Django en Excel (1 feuille par modèle)."

    def add_arguments(self, parser):
        parser.add_argument('file', nargs='?', type=str, default="Schema BD WIBChallenge.xlsx")

    def handle(self, *args, **kwargs):
        wb = Workbook()
        wb.encoding = 'utf-8'
        wb.remove(wb.active)  # Vire la feuille vide par défaut

        for model in apps.get_models():
            if model._meta.app_label not in ['accounts', 'challenges', 'questions']:
                continue

            meta = model._meta
            sheet_title = meta.verbose_name.title()[:31]
            ws = wb.create_sheet(title=sheet_title)

            # 🔹 Définir les colonnes
            headers = ["Nom Champ", "Description", "Type", "Contraintes", "Clé primaire", "Relations"]

            # 🔹 Fusion pour afficher le nom du modèle en haut (colspan)
            model_name = model.__name__
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            cell = ws.cell(row=1, column=1)
            cell.value = f"Modèle : {model_name}"
            cell.style = "Title"  # Style automatique, tu peux le personnaliser si besoin

            # 🔹 Ligne vide
            ws.append([])

            # 🔹 En-têtes des colonnes
            ws.append(headers)

            # After ws.append(headers)
            column_widths = [50, 50, 25, 50, 25, 30]  # Adjust as you like

            for i, width in enumerate(column_widths, start=1):
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = width

            # 🔹 Données des champs
            for field in meta.get_fields():
                if isinstance(field, models.ManyToOneRel) or isinstance(field, models.ManyToManyRel):
                    continue

                name = field.name
                description = getattr(field, 'verbose_name', '').encode() or ''
                type_name = field.get_internal_type()

                contraintes = []
                if hasattr(field, 'null') and not field.null:
                    contraintes.append("NOT NULL")
                if hasattr(field, 'unique') and field.unique:
                    contraintes.append("UNIQUE")
                if hasattr(field, 'blank') and not field.blank:
                    contraintes.append("REQUIRED")
                if hasattr(field, 'default') and field.default != models.fields.NOT_PROVIDED:
                    contraintes.append(
                        f"DEFAULT={field.default.__name__ if callable(field.default) else field.default}")
                contraintes_str = ', '.join(contraintes)

                is_pk = "✔" if getattr(field, 'primary_key', False) else ""

                relation = ""
                if isinstance(field, (models.ForeignKey, models.OneToOneField, models.ManyToManyField)):
                    relation = f"{field.related_model.__name__} ({field.__class__.__name__})"
                ws.append(clean_text(x) for x in [name, description, type_name, contraintes_str, is_pk, relation])

        file = kwargs['file']
        wb.save(file)
        self.stdout.write(self.style.SUCCESS(f"✅ Export terminé : {file}"))
