import os

import openpyxl
from django.core.management import BaseCommand
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.core.models import Domain


class Command(BaseCommand):
    help = 'Exporte les domaines, professions et technologies dans un fichier Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output',
            type=str,
            default='domains_professions_tools.xlsx',
            help='Chemin de sortie pour le fichier Excel'
        )

    def handle(self, *args, **options):
        output_path = options['output']

        # Création du workbook
        wb = openpyxl.Workbook()

        # Suppression de la feuille par défaut
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Récupération de tous les domaines
        domains = Domain.objects.all().prefetch_related('professions__technologies', 'professions')

        if not domains.exists():
            self.stdout.write(self.style.WARNING('Aucun domaine trouvé'))
            return

        # Pour chaque domaine, créer une feuille
        for domain in domains:
            # Création d'une nouvelle feuille pour le domaine
            sheet = wb.create_sheet(title=domain.name[:31].replace('/', ' '))  # Excel limite les noms à 31 caractères

            # Définition des styles
            title_font = Font(name='Arial', size=14, bold=True)
            header_font = Font(name='Arial', size=12, bold=True)
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            # Titre du domaine (colspan)
            sheet.merge_cells('A1:D1')
            title_cell = sheet.cell(row=1, column=1, value=domain.name)
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # En-têtes
            sheet.cell(row=3, column=1, value="Profession").font = header_font
            sheet.cell(row=3, column=1).fill = header_fill
            sheet.cell(row=3, column=2, value="Technologies").font = header_font
            sheet.cell(row=3, column=2).fill = header_fill

            # Ajustement de la hauteur des rangées
            sheet.row_dimensions[1].height = 30
            sheet.row_dimensions[3].height = 20

            # Récupération des professions du domaine
            professions = domain.professions.all()

            row_index = 4
            for profession in professions:
                # Cellule pour le nom de la profession
                prof_cell = sheet.cell(row=row_index, column=1, value=profession.title)
                prof_cell.alignment = Alignment(vertical='top', wrap_text=True)

                # Technologies de la profession
                technologies = profession.technologies.all()
                tech_names = [tech.name for tech in technologies]
                tech_text = ", ".join(tech_names) if tech_names else "Aucune technologie associée pour le moment"

                tech_cell = sheet.cell(row=row_index, column=2, value=tech_text)
                tech_cell.alignment = Alignment(vertical='top', wrap_text=True)

                # Ajustement de la hauteur de ligne en fonction du contenu
                sheet.row_dimensions[row_index].height = max(20, min(15 * (1 + tech_text.count(',')), 100))

                row_index += 1

            # Ajustement automatique de la largeur des colonnes
            for col in range(1, 3):
                column = get_column_letter(col)
                sheet.column_dimensions[column].width = 40

        # Sauvegarde du fichier
        try:
            wb.save(output_path)
            self.stdout.write(self.style.SUCCESS(
                f'Fichier Excel créé avec succès: {os.path.abspath(output_path)}'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(
                f'Erreur lors de la création du fichier Excel: {str(e)}'))
